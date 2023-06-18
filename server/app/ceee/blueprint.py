import os
import io
import zipfile

from flask import Blueprint, request, jsonify

import boto3
import requests
import xmltodict
from openpyxl import load_workbook
import pandas as pd


class S3Storage:
    def __init__(self):
        self.bucket_name = os.environ.get("BUCKET_NAME")
        self.s3_client = boto3.client("s3")

    def download_file(self, file_key: str) -> io.BytesIO:
        file_data = io.BytesIO()
        self.s3_client.download_fileobj(self, file_key, file_data)
        return file_data

    def uppload_file(self, file_key: str, file_data: io.BytesIO) -> None:
        self.s3_client.upload_fileobj(file_data, self.bucket_name, file_key)

    def get_download_url(self, file_key: str) -> str:
        return self.s3_client.generate_presigned_url(
            "get_object",
            Params={"Bucket": self.bucket_name, "Key": file_key},
            ExpiresIn=180,
        )


def formattedDate(date):
    # if formattedHour(date) != 24:
    date = (date.split("T"))[0]
    date = (
        (date.split("-"))[2] + "/" + (date.split("-"))[1] + "/" + (date.split("-"))[0]
    )
    return date
    # else:
    #    date = (date.split("T"))[0]
    #    date = datetime.strptime(date, '%Y-%m-%d') - timedelta(1)
    #    date = datetime.strftime(date, '%d/%m/%Y')
    #    return date


def formattedHeaderDate(date):
    date = (date.split("T"))[0]
    date = (date.split("-"))[1] + "/" + (date.split("-"))[0]
    return date


def formattedHour(hour):
    hour = (hour.split("T"))[1]
    hour = (hour.split("-"))[0]
    hour = (hour.split(":"))[0]
    hour = int(int(hour) + 1)
    if int(hour) == 0:
        hour = int(24)
    return hour


def formattedType(type):
    if type == "B":
        type = "Bruta"
    else:
        type = "Liquida"
    return type


def formattedOrigem(origem):
    if origem == "HCC":
        origem = "Coleta Diária"
    elif origem == "HIF":
        origem = "Hora Faltante"
    else:
        origem = "Outro"
    return origem


def postCCEE(startDate, endDate, codigoPerfil, ponto):
    url = "https://servicos.ccee.org.br:443/ws/medc/ListarMedidaBSv1"

    headers = {"Content-Type": "text/xml; charset=utf-8", "SOAPAction": "listarMedida"}

    body = (
        """<?xml version="1.0" encoding="UTF-8"?>
    <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:oas="http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-secext-1.0.xsd" xmlns:v1="http://xmlns.energia.org.br/MH/v1" xmlns:v11="http://xmlns.energia.org.br/BM/v1" xmlns:v12="http://xmlns.energia.org.br/BO/v1">
    <soapenv:Header>
        <v1:messageHeader>
            <v1:codigoPerfilAgente>"""
        + codigoPerfil
        + """</v1:codigoPerfilAgente>
        </v1:messageHeader>
        <oas:Security>
            <oas:UsernameToken>
                <oas:Username>relais</oas:Username>
                <oas:Password>08340215</oas:Password>
            </oas:UsernameToken>
        </oas:Security>
    </soapenv:Header>
    <soapenv:Body>
        <v11:listarMedida>
            <v11:pontoMedicao>
                <v12:codigo>"""
        + ponto
        + """</v12:codigo>
            </v11:pontoMedicao>
            <v11:tipoMedida>FINAL</v11:tipoMedida>
            <!--<v11:tipoMedicao>INSPECAO</v11:tipoMedicao>-->
            <v11:periodo>
                <v12:inicio>"""
        + startDate
        + """</v12:inicio>
                <v12:fim>"""
        + endDate
        + """</v12:fim>
            </v11:periodo>
        </v11:listarMedida>
    </soapenv:Body>
    </soapenv:Envelope>"""
    )

    key = os.path.dirname(os.path.realpath(__file__)) + "/PrivateKey.key"
    crt = os.path.dirname(os.path.realpath(__file__)) + "/certificadoCRT.crt"

    response = requests.post(url, data=body, headers=headers, cert=(crt, key))
    return response


def responseToDataframe(response):
    responseDict = xmltodict.parse(response.content)
    responseDict = responseDict["io2:Envelope"]["io2:Body"]["out:listarMedidaResponse"][
        "out:medidas"
    ]
    data = []
    for x in responseDict["out:medida"]:
        data.append(
            {
                "Ponto de Medição": x["out2:pontoMedicao"]["out2:codigo"],
                "Data": formattedDate(x["out2:periodo"]["out2:fim"]),
                "Hora": formattedHour(x["out2:periodo"]["out2:fim"]),
                "Tipo de Energia": formattedType(x["out2:subTipo"]),
                "Ativa Geração (kWh)": x["out2:geracaoAtiva"],
                "Ativa Consumo (kWh)": x["out2:consumoAtivo"],
                "Reativa Geração (kVArh)": x["out2:geracaoReativo"],
                "Reativa Consumo (kVArh)": x["out2:consumoReativo"],
                "Origem da Coleta": formattedOrigem(x["out2:status"]),
                "Notificação de Coleta": "",
            }
        )
    y = pd.DataFrame(data)
    y["Ativa Geração (kWh)"] = pd.to_numeric(y["Ativa Geração (kWh)"])
    y["Ativa Consumo (kWh)"] = pd.to_numeric(y["Ativa Consumo (kWh)"])
    y["Reativa Geração (kVArh)"] = pd.to_numeric(y["Reativa Geração (kVArh)"])
    y["Reativa Consumo (kVArh)"] = pd.to_numeric(y["Reativa Consumo (kVArh)"])
    return y


def setFileName(startDate, endDate, name):
    startHour = formattedHour(startDate)
    endHour = formattedHour(endDate)
    startMonth = int(((startDate.split("T"))[0]).split("-")[1])
    endMonth = int(((endDate.split("T"))[0]).split("-")[1])
    startDate = (startDate.split("T"))[0]
    startDate = "_" + (startDate.split("-"))[0] + "_" + (startDate.split("-"))[1]
    # if (startHour == 1 and endHour == 24) and ((endMonth == startMonth + 1) or (startMonth == 12 and endMonth == 1)):
    fileName = "MED_" + name + startDate
    # else:
    # fileName = "MED_" + name + startDate + "_CUSTOM"
    return fileName


ccee_bp = Blueprint("ccee", __name__)


@ccee_bp.route("/", methods=["POST"])
def ccee():
    """Download Data from Ceee

    Parameters
    ----------
    event: dict, required
        API Gateway Lambda Proxy Input Format

        Event doc: https://docs.aws.amazon.com/apigateway/latest/developerguide/set-up-lambda-proxy-integrations.html#api-gateway-simple-proxy-for-lambda-input-format

    context: object, required
        Lambda Context runtime methods and attributes

        Context doc: https://docs.aws.amazon.com/lambda/latest/dg/python-context-object.html

    Returns
    ------
    API Gateway Lambda Proxy Output Format: dict

        Return doc: https://docs.aws.amazon.com/apigateway/latest/developerguide/set-up-lambda-proxy-integrations.html
    """

    frontStringDebug = """{"startDate":"2022-08-01T01:00:00","endDate":"2022-09-01T00:00:00","tipoMedicao":"FALTANTES","infoAgente":[{"agente":"PrimaSea","codigoPerfil":"92532","agenteCCEE":"PRIMASEA","pontos":[{"ponto":"BAPRCDENTR101","descricao":"⠀","fileName":"PRIMASEA"}]},{"agente":"Américas Barra","codigoPerfil":"82044","agenteCCEE":"HOTEL AMERICAS BARRA","pontos":[{"ponto":"RJHABTENTR101","descricao":"⠀","fileName":"AMBARRA"}]},{"agente":"Américas Copacabana","codigoPerfil":"84777","agenteCCEE":"AMERICAS COPACABANA","pontos":[{"ponto":"RJACBRENTR101","descricao":"⠀","fileName":"AMCOPA"}]}],"simple":"true"}"""

    frontData = request.json
    # pd.io.formats.excel.ExcelFormatter.header_style = None
    buffer = io.BytesIO()
    zip_buffer = io.BytesIO()
    dfFaltantes = pd.DataFrame()
    for x in frontData["infoAgente"]:
        for y in x["pontos"]:
            response = postCCEE(
                frontData["startDate"],
                frontData["endDate"],
                x["codigoPerfil"],
                y["ponto"],
            )
            dframe = responseToDataframe(response)
            if frontData["tipoMedicao"] == "FINAL":
                buffer.seek(0)
                buffer.truncate(0)
                dframe.to_excel(buffer, startrow=3, index=False)
                wb = load_workbook(filename=buffer)
                ws = wb["Sheet1"]
                buffer.seek(0)
                buffer.truncate(0)
                ws["A1"] = "Tipo de Relatório: Origem de Dados da Coleta."
                ws["A2"] = "Tipo de Agente: Conectante | Nome:  " + x["agenteCCEE"]
                ws["A3"] = "Periodo Solicitado: " + formattedHeaderDate(
                    frontData["startDate"]
                )
                wb.save(buffer)
                fileName = setFileName(
                    frontData["startDate"], frontData["endDate"], y["fileName"]
                )
                if frontData["simple"] == "true":
                    with zipfile.ZipFile(
                        zip_buffer, "a", zipfile.ZIP_DEFLATED, False
                    ) as zip_file:
                        zip_file.writestr(fileName + ".xlsx", buffer.getvalue())
            else:
                dframe = dframe[
                    (dframe["Origem da Coleta"] == "Hora Faltante")
                    | (dframe["Origem da Coleta"] == "Outro")
                ]
                dframe = (
                    dframe.groupby("Data")["Hora"].apply(list).reset_index(name="Horas")
                )
                dframe.insert(0, "Arquivo", y["fileName"])
                dfFaltantes = pd.concat([dfFaltantes, dframe])

    if frontData["tipoMedicao"] == "FALTANTES":
        writer = pd.ExcelWriter(buffer, engine="xlsxwriter")
        dfFaltantes.to_excel(writer, startrow=1, index=False, header=False)
        wb = writer.book
        ws = writer.sheets["Sheet1"]
        (max_row, max_col) = dfFaltantes.shape
        column_settings = []
        for header in dfFaltantes.columns:
            column_settings.append({"header": header})
        ws.add_table(0, 0, max_row, max_col - 1, {"columns": column_settings})
        ws.set_column(0, max_col - 1, 25)
        ws.set_column(max_col - 1, max_col - 1, 65)
        buffer.seek(0)
        buffer.truncate(0)
        writer.close()
        buffer.seek(0)
        fileName = (
            "HORAS FALTANTES - " + (frontData["startDate"].split("T"))[0] + ".xlsx"
        )
        s3_client = S3Storage()
        s3_client.uppload_file(fileName, buffer)
        return (jsonify({"download_url": s3_client.get_download_url(fileName)}), 200)

    if frontData["tipoMedicao"] == "FINAL":
        if frontData["simple"] == "true":
            zip_buffer.seek(0)
            s3_client = S3Storage()
            s3_client.uppload_file(fileName + ".zip", zip_buffer)
            return (
                jsonify(
                    {"download_url": s3_client.get_download_url(fileName + ".zip")}
                ),
                200,
            )

        s3_client = S3Storage()
        buffer.seek(0)
        s3_client.uppload_file(fileName + ".xlsx", buffer)
        return (
            jsonify({"download_url": s3_client.get_download_url(fileName + ".xlsx")}),
            200,
        )

    return (jsonify({"message": "download failed"}), 200)
